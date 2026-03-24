import face_recognition
import cv2
import os
from datetime import datetime, timedelta
import openpyxl
import pandas as pd  # This library is crucial for the final CSV export

# --- 1. SETUP: Create necessary folders ---
# This ensures the folder for saving face snapshots exists.
os.makedirs("captured_faces", exist_ok=True)


# --- 2. LOAD KNOWN FACES: Teach the system who to recognize ---
print("[INFO] Loading known faces...")
known_face_encodings = []
known_face_names = []
known_faces_dir = "known_faces"

# Loop through each image in the 'known_faces' folder
for filename in os.listdir(known_faces_dir):
    if filename.lower().endswith((".jpg", ".jpeg", ".png")):
        image_path = os.path.join(known_faces_dir, filename)
        # Load the image
        image = face_recognition.load_image_file(image_path)
        # Find the face encoding (the unique mathematical signature)
        encodings = face_recognition.face_encodings(image)
        
        if encodings:
            known_face_encodings.append(encodings[0])
            # Use the filename (without extension) as the person's name
            known_face_names.append(os.path.splitext(filename)[0])
        else:
            print(f"[WARNING] No face found in {filename}. Please use a clearer picture.")


# --- 3. PREPARE LOG FILES: Set up the Excel and CSV files ---
log_file = "face_log.xlsx"
today_date = datetime.now().strftime("%Y-%m-%d")
today_csv = f"face_log_{today_date}.csv"

# Create the Excel file if it doesn't exist
if not os.path.exists(log_file):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = today_date
    ws.append(["Name", "Date", "Time", "Status"])
    wb.save(log_file)

wb = openpyxl.load_workbook(log_file)

# Create a new sheet for today's date if it doesn't exist
if today_date not in wb.sheetnames:
    ws = wb.create_sheet(title=today_date)
    ws.append(["Name", "Date", "Time", "Status"])
else:
    ws = wb[today_date]


# --- 4. INITIALIZE VARIABLES: Prepare for real-time tracking ---
last_seen_times = {}
status = {}
summary = {"Login": 0, "Logout": 0}
login_row_index = {}


# --- 5. START WEBCAM: The main recognition loop ---
video_capture = cv2.VideoCapture(0)
print("[INFO] Starting face recognition. Press 'q' on the camera window to quit.")

while True:
    # Grab a single frame of video
    ret, frame = video_capture.read()
    if not ret:
        print("[ERROR] Failed to grab frame from camera.")
        break

    # Resize frame for faster processing (to 1/4 size)
    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
    # Convert the image from BGR color (which OpenCV uses) to RGB color
    rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

    # Find all the faces and face encodings in the current frame of video
    face_locations = face_recognition.face_locations(rgb_small_frame)
    face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")

    # Loop through each face found in the frame
    for face_encoding, face_location in zip(face_encodings, face_locations):
        # See if the face is a match for the known face(s)
        face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
        best_match_index = face_distances.argmin() if len(face_distances) > 0 else None

        name = "Unknown"

        # If a match was found, use a tolerance of 0.45 for accuracy
        if best_match_index is not None and face_distances[best_match_index] < 0.45:
            name = known_face_names[best_match_index]

            # --- ATTENDANCE LOGIC ---
            last_seen = last_seen_times.get(name)
            state = status.get(name)
 
            # Scenario 1: First time seeing the person today, or they are returning after a logout
            if not last_seen or state == "Logout":
                ws.append([name, date_str, time_str, "Login"])
                login_row_index[name] = ws.max_row
                status[name] = "Login"
                summary["Login"] += 1

                # Save a snapshot for verification
                top, right, bottom, left = [v * 4 for v in face_location]
                face_image = frame[top:bottom, left:right]
                face_dir = os.path.join("captured_faces", date_str)
                os.makedirs(face_dir, exist_ok=True)
                face_filename = os.path.join(face_dir, f"{name}_Login_{now.strftime('%H%M%S')}.jpg")
                cv2.imwrite(face_filename, face_image)

            # Scenario 2: Person returns after a long break (more than 30 minutes)
            elif (now - last_seen) > timedelta(minutes=30):
                # First, log them out
                ws.append([name, date_str, time_str, "Logout"])
                summary["Logout"] += 1
                status[name] = "Logout"
                
                # Save logout snapshot
                top, right, bottom, left = [v * 4 for v in face_location]
                face_image_logout = frame[top:bottom, left:right]
                face_dir = os.path.join("captured_faces", date_str)
                os.makedirs(face_dir, exist_ok=True)
                face_filename_logout = os.path.join(face_dir, f"{name}_Logout_{now.strftime('%H%M%S')}.jpg")
                cv2.imwrite(face_filename_logout, face_image_logout)

                # Then, immediately log them in for their new session
                ws.append([name, date_str, time_str, "Login"])
                login_row_index[name] = ws.max_row
                status[name] = "Login"
                summary["Login"] += 1
                
                # Save login snapshot
                face_image_login = frame[top:bottom, left:right]
                face_filename_login = os.path.join(face_dir, f"{name}_Login_{now.strftime('%H%M%S')}.jpg")
                cv2.imwrite(face_filename_login, face_image_login)

            # Scenario 3: Person is still present (seen again within 30 minutes)
            elif state == "Login":
                # To keep the log "live", we can optionally update the time of the last login entry
                # This shows the last time they were "seen" present.
                row_idx = login_row_index.get(name)
                if row_idx and row_idx <= ws.max_row:
                    ws.cell(row=row_idx, column=3).value = time_str

            # Always update the last seen time for the person
            last_seen_times[name] = now

        # --- DRAW VISUALS: Display the results on the screen ---
        top, right, bottom, left = [v * 4 for v in face_location]
        # Draw a box around the face
        cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
        # Draw a label with a name below the face
        cv2.putText(frame, name, (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 255, 255), 2)

    # Display the resulting image
    cv2.imshow("Student Face Recognition", frame)

    # Hit 'q' on the keyboard to quit!
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break


# --- 6. CLEANUP: Release handle to the webcam and save files ---
video_capture.release()
cv2.destroyAllWindows()
wb.save(log_file)

# Print a summary to the console
print("\n[SUMMARY]")
print(f"Total Logins:  {summary['Login']}")
print(f"Total Logouts: {summary['Logout']}")

# --- 7. FINAL EXPORT: Save today's log to a clean CSV for the dashboard ---
print(f"[INFO] Saving data to CSV: {today_csv}")
ws = wb[today_date]
# Convert the Excel sheet to a pandas DataFrame for easy saving
df = pd.DataFrame(ws.values)

if not df.empty:
    # Set the first row as the column headers
    df.columns = df.iloc[0]
    df = df[1:]
    # Save to CSV without the index column
    df.to_csv(today_csv, index=False)
    print(f"[INFO] Data successfully exported to: {today_csv}")
else:
    print("[INFO] No data to export for today.")