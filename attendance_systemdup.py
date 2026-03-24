import face_recognition
import cv2
import os
from datetime import datetime, timedelta
import openpyxl
import csv

os.makedirs("captured_faces", exist_ok=True)

known_face_encodings = []
known_face_names = []
known_faces_dir = "known_faces"

for filename in os.listdir(known_faces_dir):
    if filename.lower().endswith((".jpg", ".jpeg", ".png")):
        image_path = os.path.join(known_faces_dir, filename)
        image = face_recognition.load_image_file(image_path)
        encoding = face_recognition.face_encodings(image)
        if encoding:
            known_face_encodings.append(encoding[0])
            known_face_names.append(os.path.splitext(filename)[0])
        else:
            print(f"[WARNING] No face found in {filename}")

log_file = "face_log.xlsx"
today_date = datetime.now().strftime("%Y-%m-%d")
today_csv = f"face_log_{today_date}.csv"

if not os.path.exists(log_file):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = today_date
    ws.append(["Name", "Date", "Time", "Status"])
    wb.save(log_file)

wb = openpyxl.load_workbook(log_file)

if today_date in wb.sheetnames:
    ws = wb[today_date]
else:
    ws = wb.create_sheet(title=today_date)
    ws.append(["Name", "Date", "Time", "Status"])

last_seen_times = {}
status = {}
summary = {"Login": 0, "Logout": 0}
login_row_index = {}

video_capture = cv2.VideoCapture(0)
print("[INFO] Starting face recognition. Press 'q' to quit.")

while True:
    ret, frame = video_capture.read()
    if not ret:
        break

    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
    rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

    face_locations = face_recognition.face_locations(rgb_small_frame)
    face_encodings = face_recognition.face_encodings(
        rgb_small_frame, face_locations)

    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")

    for face_encoding, face_location in zip(face_encodings, face_locations):
        face_distances = face_recognition.face_distance(
            known_face_encodings, face_encoding)
        best_match_index = face_distances.argmin(
        ) if len(face_distances) > 0 else None

        name = "Unknown"

        if best_match_index is not None and face_distances[best_match_index] < 0.45:
            name = known_face_names[best_match_index]

            top, right, bottom, left = [v * 4 for v in face_location]
            face_width = right - left
            face_height = bottom - top

            if face_width < 60 or face_height < 60:
                continue

            last_seen = last_seen_times.get(name)
            state = status.get(name)

            if not last_seen or state == "Logout":
                ws.append([name, date_str, time_str, "Login"])
                login_row_index[name] = ws.max_row
                status[name] = "Login"
                summary["Login"] += 1

                face_image = frame[top:bottom, left:right]
                face_dir = os.path.join("captured_faces", today_date)
                os.makedirs(face_dir, exist_ok=True)
                face_filename = os.path.join(
                    face_dir, f"{name}_Login_{now.strftime('%H%M%S')}.jpg")
                cv2.imwrite(face_filename, face_image)

            elif (now - last_seen) > timedelta(minutes=30):
                ws.append([name, date_str, time_str, "Logout"])
                summary["Logout"] += 1
                status[name] = "Logout"

                face_image = frame[top:bottom, left:right]
                face_dir = os.path.join("captured_faces", today_date)
                os.makedirs(face_dir, exist_ok=True)
                face_filename = os.path.join(
                    face_dir, f"{name}_Logout_{now.strftime('%H%M%S')}.jpg")
                cv2.imwrite(face_filename, face_image)

                wb.save(log_file)
                continue

            elif state == "Login":
                row_idx = login_row_index.get(name)
                if row_idx and row_idx <= ws.max_row:
                    ws.cell(row=row_idx, column=3).value = time_str
                    wb.save(log_file)

            last_seen_times[name] = now

        top, right, bottom, left = [v * 4 for v in face_location]
        cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
        cv2.putText(frame, name, (left, top - 10),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 255, 255), 2)

    cv2.imshow("Student Face Recognition", frame)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

video_capture.release()
cv2.destroyAllWindows()
wb.save(log_file)

print("\n[SUMMARY]")
print(f"Total Logins:  {summary['Login']}")
print(f"Total Logouts: {summary['Logout']}")

ws = wb[today_date]
with open(today_csv, mode='w', newline='') as file:
    writer = csv.writer(file)
    for row in ws.iter_rows(values_only=True):
        writer.writerow(row)

print(f"[INFO] Data exported to: {today_csv}")